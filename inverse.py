

from openpyxl import Workbook


# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Inverse Table"



# Set headers
ws.cell(row=1, column=1).value = "Number"
ws.cell(row=1, column=2).value = "Inverse"




# Calculate and write inverses
for i in range(2, 101):
  number = i - 1
  if number != 0:
    inverse = 1 / number
  else:
    inverse = "Dividing by Zero Error"
  ws.cell(row=i, column=1).value = number
  ws.cell(row=i, column=2).value = inverse

# Save the workbook
wb.save("inverse_table.xlsx")

print("Inverse table created successfully in 'inverse_table.xlsx' file.")

